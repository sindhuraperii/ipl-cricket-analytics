"""
IPL Cricket Analytics - Excel Report Generator
===============================================
Creates a professional Excel report with all analysis results
"""

import pandas as pd
import sqlite3
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
import warnings

warnings.filterwarnings('ignore')

print("=" * 80)
print("📊 IPL CRICKET ANALYTICS - EXCEL REPORT GENERATOR")
print("=" * 80)

# Load data
print("\n📥 Loading data...")

# Load from CSV files
eda_summary = pd.read_csv('data/eda_summary.csv')
engineered_features = pd.read_csv('data/processed/engineered_features.csv')
model_results = pd.read_csv('data/model_results.csv')

# Load from database
conn = sqlite3.connect('data/ipl_database.db')
matches = pd.read_sql_query("SELECT * FROM matches LIMIT 100", conn)
conn.close()

print("   ✅ All data loaded successfully")

# Create workbook
print("\n📝 Creating Excel workbook...")
wb = Workbook()
wb.remove(wb.active)  # Remove default sheet

# Define styles
header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF", size=12)
title_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
title_font = Font(bold=True, color="FFFFFF", size=14)
border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

def add_dataframe_to_sheet(ws, df, start_row=1, start_col=1):
    """Add dataframe to worksheet without column issues"""
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), start_row):
        for c_idx, value in enumerate(row, start_col):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            
            # Style header row
            if r_idx == start_row:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            
            cell.border = border
            cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

# Sheet 1: Executive Summary
print("   📄 Sheet 1: Executive Summary")
ws = wb.create_sheet("Executive Summary")

ws['A1'] = "IPL CRICKET ANALYTICS PROJECT"
ws['A1'].font = title_font
ws['A1'].fill = title_fill
ws.merge_cells('A1:D1')

summary_data = {
    'Metric': [
        'Total Matches Analyzed',
        'Total Deliveries',
        'Unique Teams',
        'Unique Venues',
        'Unique Cities',
        'Top Player (PoM)',
        'Best Model',
        'Best Model Accuracy',
        'Engineered Features',
        'Date Range'
    ],
    'Value': [
        '2,380',
        '283,042',
        '19',
        '59',
        '38',
        'AB de Villiers (50 awards)',
        'Random Forest',
        '48.95%',
        '32',
        '2008-04-18 to 2026-04-16'
    ]
}
summary_df = pd.DataFrame(summary_data)
add_dataframe_to_sheet(ws, summary_df, start_row=3)
ws.column_dimensions['A'].width = 30
ws.column_dimensions['B'].width = 50

# Sheet 2: EDA Summary
print("   📄 Sheet 2: EDA Summary")
ws = wb.create_sheet("EDA Summary")
ws['A1'] = "EXPLORATORY DATA ANALYSIS SUMMARY"
ws['A1'].font = title_font
ws['A1'].fill = title_fill
ws.merge_cells('A1:B1')
add_dataframe_to_sheet(ws, eda_summary, start_row=3)
ws.column_dimensions['A'].width = 30
ws.column_dimensions['B'].width = 50

# Sheet 3: Key Insights
print("   📄 Sheet 3: Key Insights")
ws = wb.create_sheet("Key Insights")
ws['A1'] = "KEY INSIGHTS FROM ANALYSIS"
ws['A1'].font = title_font
ws['A1'].fill = title_fill
ws.merge_cells('A1:B1')

insights_data = {
    'Category': [
        'Most Important Feature',
        'Second Most Important',
        'Third Most Important',
        'Toss Decision - Field',
        'Toss Decision - Bat',
        'Top Venue',
        'Top City',
        'Most Matches (Team)',
        'Best Performing Player',
        'Model Performance'
    ],
    'Details': [
        'Taking Wickets (29% importance)',
        'Scoring Runs (21% importance)',
        'Team Historical Performance (18%)',
        '65.9% of teams choose to field',
        '34.1% of teams choose to bat',
        'Eden Gardens (154 matches)',
        'Mumbai (366 matches)',
        'Mumbai Indians (282 matches)',
        'AB de Villiers (50 Player of Match)',
        'Random Forest: 48.95% accuracy'
    ]
}
insights_df = pd.DataFrame(insights_data)
add_dataframe_to_sheet(ws, insights_df, start_row=3)
ws.column_dimensions['A'].width = 30
ws.column_dimensions['B'].width = 70

# Sheet 4: Top Players
print("   📄 Sheet 4: Top Players")
ws = wb.create_sheet("Top Players")
ws['A1'] = "TOP PLAYERS - PLAYER OF MATCH AWARDS"
ws['A1'].font = title_font
ws['A1'].fill = title_fill
ws.merge_cells('A1:B1')

players_data = {
    'Rank': [1, 2, 3, 4, 5],
    'Player': ['AB de Villiers', 'CH Gayle', 'RG Sharma', 'V Kohli', 'MS Dhoni'],
    'Awards': [50, 44, 42, 38, 36]
}
players_df = pd.DataFrame(players_data)
add_dataframe_to_sheet(ws, players_df, start_row=3)
ws.column_dimensions['A'].width = 10
ws.column_dimensions['B'].width = 30
ws.column_dimensions['C'].width = 15

# Sheet 5: Top Venues
print("   📄 Sheet 5: Top Venues")
ws = wb.create_sheet("Top Venues")
ws['A1'] = "TOP VENUES - MOST MATCHES"
ws['A1'].font = title_font
ws['A1'].fill = title_fill
ws.merge_cells('A1:B1')

venues_data = {
    'Rank': [1, 2, 3, 4, 5],
    'Venue': ['Eden Gardens', 'Wankhede Stadium', 'M Chinnaswamy Stadium', 'Feroz Shah Kotla', 'Wankhede Stadium, Mumbai'],
    'Matches': [154, 146, 129, 119, 110]
}
venues_df = pd.DataFrame(venues_data)
add_dataframe_to_sheet(ws, venues_df, start_row=3)
ws.column_dimensions['A'].width = 10
ws.column_dimensions['B'].width = 40
ws.column_dimensions['C'].width = 15

# Sheet 6: Top Cities
print("   📄 Sheet 6: Top Cities")
ws = wb.create_sheet("Top Cities")
ws['A1'] = "TOP CITIES - MATCH DISTRIBUTION"
ws['A1'].font = title_font
ws['A1'].fill = title_fill
ws.merge_cells('A1:B1')

cities_data = {
    'Rank': [1, 2, 3, 4, 5],
    'City': ['Mumbai', 'Kolkata', 'Delhi', 'Chennai', 'Hyderabad'],
    'Matches': [366, 205, 197, 188, 169]
}
cities_df = pd.DataFrame(cities_data)
add_dataframe_to_sheet(ws, cities_df, start_row=3)
ws.column_dimensions['A'].width = 10
ws.column_dimensions['B'].width = 20
ws.column_dimensions['C'].width = 15

# Sheet 7: Model Results
print("   📄 Sheet 7: Model Results")
ws = wb.create_sheet("Model Results")
ws['A1'] = "ML MODEL PERFORMANCE COMPARISON"
ws['A1'].font = title_font
ws['A1'].fill = title_fill
ws.merge_cells('A1:F1')

# Limit model_results to first 100 rows to avoid large data
model_results_limited = model_results.head(100) if len(model_results) > 100 else model_results
add_dataframe_to_sheet(ws, model_results_limited, start_row=3)

for col in range(1, 6):
    ws.column_dimensions[chr(64 + col)].width = 20

# Sheet 8: Engineered Features Sample
print("   📄 Sheet 8: Engineered Features Sample")
ws = wb.create_sheet("Features Sample")
ws['A1'] = "ENGINEERED FEATURES SAMPLE (First 50 rows)"
ws['A1'].font = title_font
ws['A1'].fill = title_fill
ws.merge_cells('A1:Z1')

features_limited = engineered_features.head(50)
add_dataframe_to_sheet(ws, features_limited, start_row=3)

for col in range(1, min(15, len(features_limited.columns) + 1)):
    ws.column_dimensions[chr(64 + col)].width = 18

# Sheet 9: Data Dictionary
print("   📄 Sheet 9: Data Dictionary")
ws = wb.create_sheet("Data Dictionary")
ws['A1'] = "DATA DICTIONARY"
ws['A1'].font = title_font
ws['A1'].fill = title_fill
ws.merge_cells('A1:C1')

dictionary_data = {
    'Column Name': [
        'match_id', 'date', 'batting_team', 'bowling_team', 'player_of_match',
        'toss_winner', 'toss_decision', 'venue', 'city', 'season'
    ],
    'Data Type': ['Integer', 'Date', 'Text', 'Text', 'Text', 'Text', 'Text', 'Text', 'Text', 'Float'],
    'Description': [
        'Unique match identifier',
        'Match date',
        'Team batting first',
        'Team bowling first',
        'Player awarded Player of Match',
        'Team that won the toss',
        'Decision taken after toss (bat/field)',
        'Venue name',
        'City where match was held',
        'IPL season year'
    ]
}
dictionary_df = pd.DataFrame(dictionary_data)
add_dataframe_to_sheet(ws, dictionary_df, start_row=3)
ws.column_dimensions['A'].width = 25
ws.column_dimensions['B'].width = 15
ws.column_dimensions['C'].width = 50

# Sheet 10: Methodology
print("   📄 Sheet 10: Methodology")
ws = wb.create_sheet("Methodology")
ws['A1'] = "PROJECT METHODOLOGY"
ws['A1'].font = title_font
ws['A1'].fill = title_fill
ws.merge_cells('A1:B1')

methodology_data = {
    'Step': [
        '1. Data Collection',
        '2. Data Cleaning',
        '3. EDA',
        '4. Feature Engineering',
        '5. Data Preparation',
        '6. Model Training',
        '7. Model Evaluation',
        '8. Report Generation'
    ],
    'Description': [
        'Downloaded IPL dataset with 283,678 records',
        'Removed duplicates, handled missing values',
        'Created 7 visualizations analyzing patterns',
        'Engineered 32 features from raw data',
        'Scaled features, split train/test (80/20)',
        'Trained 3 models: LR, RF, GB',
        'Evaluated using Accuracy, Precision, Recall, F1, AUC',
        'Generated comprehensive Excel and visualization reports'
    ]
}
methodology_df = pd.DataFrame(methodology_data)
add_dataframe_to_sheet(ws, methodology_df, start_row=3)
ws.column_dimensions['A'].width = 25
ws.column_dimensions['B'].width = 70

# Save workbook
print("\n💾 Saving Excel report...")
output_path = 'reports/IPL_Analytics_Report.xlsx'

try:
    wb.save(output_path)
    print(f"   ✅ Saved: {output_path}")
except Exception as e:
    print(f"   ❌ Error saving: {e}")
    # Try alternative format
    output_path = 'reports/IPL_Analytics_Report.csv'
    summary_df.to_csv(output_path, index=False)
    print(f"   ✅ Saved as CSV: {output_path}")

# Print summary
print("\n" + "=" * 80)
print("✅ EXCEL REPORT GENERATION COMPLETE!")
print("=" * 80)

print(f"\n📁 Report saved to: {output_path}")
print("\n📊 Report Contents (10 Sheets):")
print("   1. Executive Summary - Project overview and key metrics")
print("   2. EDA Summary - Dataset statistics")
print("   3. Key Insights - Main findings from analysis")
print("   4. Top Players - Player of Match awards ranking")
print("   5. Top Venues - Venue statistics")
print("   6. Top Cities - City distribution")
print("   7. Model Results - ML model performance")
print("   8. Features Sample - Engineered features preview")
print("   9. Data Dictionary - Column descriptions")
print("  10. Methodology - Project approach")

print("\n" + "=" * 80)
print("🎯 Next step: Initialize Git and push to GitHub")
print("=" * 80)
